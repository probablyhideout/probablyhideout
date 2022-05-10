import requests
import openpyxl

res = requests.get('https://zipcloud.ibsnet.co.jp/api/search?zipcode=1000001')
print(res.status_code)
print(res.text)
res_json = res.json()
results = res_json['results'][0]
address = results['address1'] + results['address2'] + results['address3']
print(address)


wb = openpyxl.load_workbook('./test_webexcel.xlsx')
sheet = wb['Sheet1']
sheet['A1'].value = results['address1']
sheet['B1'].value = results['address2']
sheet['C1'].value = results['address3']
wb.save('./test_webexcel.xlsx')
wb.close()

res = requests.get('https://zipcloud.ibsnet.co.jp/api/search?zipcode=1000002')
print(res.status_code)
print(res.text)
res_json = res.json()
results = res_json['results'][0]
address = results['address1'] + results['address2'] + results['address3']
print(address)


wb = openpyxl.load_workbook('./test_webexcel.xlsx')
sheet = wb['Sheet1']
sheet['A2'].value = results['address1']
sheet['B2'].value = results['address2']
sheet['C2'].value = results['address3']
wb.save('./test_webexcel.xlsx')
wb.close()

res = requests.get('https://zipcloud.ibsnet.co.jp/api/search?zipcode=1000003')
print(res.status_code)
print(res.text)
res_json = res.json()
results = res_json['results'][0]
address = results['address1'] + results['address2'] + results['address3']
print(address)

import openpyxl
wb = openpyxl.load_workbook('./test_webexcel.xlsx')
sheet = wb['Sheet1']
sheet['A3'].value = results['address1']
sheet['B3'].value = results['address2']
sheet['C3'].value = results['address3']
wb.save('./test_webexcel.xlsx')
wb.close()